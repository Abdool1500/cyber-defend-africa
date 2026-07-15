"""Branded Excel report generation using openpyxl. Same philosophy as
apps.reports.services.pdf: small, focused builder functions per report
type rather than one generic "headers + rows" engine — each report
shapes its own data (and, for Feedback in particular, its own privacy
rules), so a shared abstraction would just hide those differences.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _style_header_row(ws, headers):
    header_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_index)].width = 20


def _new_sheet(title, headers):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    _style_header_row(ws, headers)
    return wb, ws


def _to_bytes(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_impact_report_excel(impact) -> bytes:
    wb, ws = _new_sheet("Impact Report", ["Metric", "Value", "Type"])
    computed_labels = {
        "people_trained": "People Trained",
        "certificates_issued": "Certificates Issued",
        "total_learning_hours": "Total Learning Hours",
        "labs_completed": "Labs Completed",
        "completion_rate": "Completion Rate (%)",
        "average_score": "Average Quiz Score (%)",
        "avg_skill_improvement": "Average Skill Improvement (%)",
        "employment_rate": "Employment Rate (%)",
        "nps_score": "Net Promoter Score",
    }
    for key, label in computed_labels.items():
        ws.append([label, impact["computed"][key], "Computed"])

    self_reported_labels = {
        "smes_protected": "SMEs Protected",
        "healthcare_workers_trained": "Healthcare Workers Trained",
        "businesses_started": "Businesses Started",
    }
    for key, label in self_reported_labels.items():
        ws.append([label, impact["self_reported"][key], "Self-Reported"])
    as_of = impact["self_reported"]["as_of"]
    if as_of:
        ws.append(["Self-Reported As Of", as_of.strftime("%Y-%m-%d"), "Self-Reported"])
    return _to_bytes(wb)


def build_course_report_excel(rows) -> bytes:
    wb, ws = _new_sheet("Course Report", [
        "Course", "Total Enrollments", "Completed", "Completion Rate (%)",
        "Average Score (%)", "Average Rating", "Certificates Issued",
    ])
    for row in rows:
        ws.append([
            row["course_title"], row["total_enrollments"], row["completed_enrollments"],
            row["completion_rate"], row["average_score"], row["average_rating"],
            row["certificates_issued"],
        ])
    return _to_bytes(wb)


def build_student_report_excel(rows) -> bytes:
    wb, ws = _new_sheet("Student Report", [
        "Student", "Email", "Total Enrollments", "Completed",
        "Average Score (%)", "Certificates Earned", "Learning Hours",
    ])
    for row in rows:
        ws.append([
            row["student_name"], row["student_email"], row["total_enrollments"],
            row["completed_enrollments"], row["average_score"], row["certificates_earned"],
            row["learning_hours"],
        ])
    return _to_bytes(wb)


def build_employment_report_excel(employment) -> bytes:
    wb, ws = _new_sheet("Employment Report", ["Status", "Count"])
    for row in employment["status_breakdown"]:
        ws.append([row["label"], row["count"]])
    ws.append(["Total Reporting", employment["total_reporting"]])
    ws.append(["Employment Rate (%)", employment["employment_rate"]])
    return _to_bytes(wb)


def _add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    _style_header_row(ws, headers)
    for row in rows:
        ws.append(row)


def build_fid_impact_report_excel(data) -> bytes:
    """One workbook, one sheet per report section — a funder can flip
    between tabs rather than scroll through one long flat sheet.
    Mirrors the section grouping in build_fid_impact_report_pdf()."""
    wb = Workbook()
    wb.remove(wb.active)  # the default blank sheet — every sheet here is explicitly added

    reach = data["reach"]
    _add_sheet(wb, "Reach", ["Metric", "Value"], [
        ["People Trained", reach["people_trained"]],
        ["Certificates Issued", reach["certificates_issued"]],
        ["Total Learning Hours", reach["total_learning_hours"]],
        ["Labs Completed", reach["labs_completed"]],
    ])

    outcomes = data["learning_outcomes"]
    _add_sheet(wb, "Learning Outcomes", ["Metric", "Value"], [
        ["Completion Rate (%)", outcomes["completion_rate"]],
        ["Average Quiz Score (%)", outcomes["average_score"]],
        ["Average Skill Improvement (%)", outcomes["avg_skill_improvement"]],
        ["Net Promoter Score", outcomes["nps_score"]],
    ])

    employment = data["employment_outcomes"]
    employment_rows = [[row["label"], row["count"]] for row in employment["status_breakdown"]]
    employment_rows.append(["Total Reporting", employment["total_reporting"]])
    employment_rows.append(["Employment Rate (%)", employment["employment_rate"]])
    _add_sheet(wb, "Employment Outcomes", ["Status", "Count"], employment_rows)

    field = data["field_impact"]
    field_rows = [
        ["SMEs Protected", field["smes_protected"]],
        ["Healthcare Workers Trained", field["healthcare_workers_trained"]],
        ["Businesses Started", field["businesses_started"]],
    ]
    if field["as_of"]:
        field_rows.append(["As Of", field["as_of"].strftime("%Y-%m-%d")])
    _add_sheet(wb, "Field Impact", ["Metric", "Value"], field_rows)

    if data["cohorts"]:
        _add_sheet(
            wb, "Cohort Breakdown",
            ["Cohort", "Members", "Completion %", "Avg Score", "Certificates"],
            [
                [c["name"], c["member_count"], c["completion_rate"], c["average_score"], c["certificates_earned"]]
                for c in data["cohorts"]
            ],
        )

    if data["top_courses"]:
        _add_sheet(
            wb, "Top Courses", ["Course", "Enrollments"],
            [[row["course__title"], row["enrollment_count"]] for row in data["top_courses"]],
        )

    return _to_bytes(wb)
