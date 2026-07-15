from django.test import TestCase
from django.urls import reverse

from apps.accounts.models import User


class AnalyticsDashboardTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            email="reportsadmin@example.com", password="pass1234", full_name="Reports Admin",
            role=User.Role.ADMIN, status=User.Status.ACTIVE, is_staff=True,
        )
        self.client.login(username="reportsadmin@example.com", password="pass1234")

    def test_dashboard_renders_with_zero_state(self):
        response = self.client.get(reverse("reports:analytics_dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Analytics Dashboard")

    def test_student_cannot_access_dashboard(self):
        self.client.logout()
        User.objects.create_user(
            email="reportsstudent@example.com", password="pass1234", full_name="Student",
            role=User.Role.STUDENT, status=User.Status.ACTIVE,
        )
        self.client.login(username="reportsstudent@example.com", password="pass1234")
        response = self.client.get(reverse("reports:analytics_dashboard"))
        self.assertEqual(response.status_code, 403)

    def test_instructor_cannot_access_dashboard(self):
        self.client.logout()
        User.objects.create_user(
            email="reportsinstructor@example.com", password="pass1234", full_name="Instructor",
            role=User.Role.INSTRUCTOR, status=User.Status.ACTIVE,
        )
        self.client.login(username="reportsinstructor@example.com", password="pass1234")
        response = self.client.get(reverse("reports:analytics_dashboard"))
        self.assertEqual(response.status_code, 403)


class ImpactDashboardTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            email="impactadmin@example.com", password="pass1234", full_name="Impact Admin",
            role=User.Role.ADMIN, status=User.Status.ACTIVE, is_staff=True,
        )
        self.client.login(username="impactadmin@example.com", password="pass1234")

    def test_dashboard_renders_with_zero_state(self):
        response = self.client.get(reverse("reports:impact_dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Impact Dashboard")
        self.assertContains(response, "No figures recorded yet.")

    def test_dashboard_shows_self_reported_snapshot(self):
        from apps.analytics.models import ImpactSnapshot

        ImpactSnapshot.objects.create(smes_protected=42, healthcare_workers_trained=7, businesses_started=3)
        response = self.client.get(reverse("reports:impact_dashboard"))
        self.assertContains(response, "42")
        self.assertContains(response, "7")
        self.assertContains(response, "3")

    def test_export_csv_contains_computed_and_self_reported_rows(self):
        from apps.analytics.models import ImpactSnapshot

        ImpactSnapshot.objects.create(smes_protected=5, healthcare_workers_trained=2, businesses_started=1)
        response = self.client.get(reverse("reports:export_impact_csv"))
        content = response.content.decode()
        self.assertIn("People Trained", content)
        self.assertIn("SMEs Protected", content)
        self.assertIn("Computed", content)
        self.assertIn("Self-Reported", content)

    def test_student_cannot_access_dashboard(self):
        self.client.logout()
        User.objects.create_user(
            email="impactstudentuser@example.com", password="pass1234", full_name="Student",
            role=User.Role.STUDENT, status=User.Status.ACTIVE,
        )
        self.client.login(username="impactstudentuser@example.com", password="pass1234")
        response = self.client.get(reverse("reports:impact_dashboard"))
        self.assertEqual(response.status_code, 403)


class ExportEndpointsTests(TestCase):
    """16.L — every new CSV/Excel/PDF export route, exercised against
    real fixture data (not just an empty platform) so the openpyxl/
    ReportLab builders actually run their row-writing code paths, not
    just their empty-state branches."""

    def setUp(self):
        from apps.certificates.models import Certificate
        from apps.courses.models import Course
        from apps.employment.models import EmploymentOutcome
        from apps.enrollments.models import Enrollment
        from apps.quizzes.models import Quiz, QuizAttempt

        self.admin = User.objects.create_user(
            email="exportadmin@example.com", password="pass1234", full_name="Export Admin",
            role=User.Role.ADMIN, status=User.Status.ACTIVE, is_staff=True,
        )
        self.client.login(username="exportadmin@example.com", password="pass1234")

        self.student = User.objects.create_user(
            email="exportstudent@example.com", password="pass1234", full_name="Export Student",
            role=User.Role.STUDENT, status=User.Status.ACTIVE,
        )
        self.course = Course.objects.create(
            title="Export Course", slug="export-course", short_description="x", description="x",
            status=Course.Status.PUBLISHED,
        )
        Enrollment.objects.create(student=self.student, course=self.course, status=Enrollment.Status.COMPLETED)
        quiz = Quiz.objects.create(course=self.course, title="Q1", status=Quiz.Status.PUBLISHED)
        QuizAttempt.objects.create(
            quiz=quiz, student=self.student, attempt_number=1, status=QuizAttempt.Status.GRADED,
            percentage=88.0, score=88, max_score=100,
        )
        Certificate.objects.create(student=self.student, course=self.course, status=Certificate.Status.ISSUED)
        EmploymentOutcome.objects.create(student=self.student, status=EmploymentOutcome.Status.EMPLOYED_FULL_TIME)

    def _assert_ok(self, url_name, content_type):
        response = self.client.get(reverse(f"reports:{url_name}"))
        self.assertEqual(response.status_code, 200, url_name)
        self.assertEqual(response["Content-Type"], content_type)
        self.assertGreater(len(response.content), 0, url_name)

    def test_impact_exports(self):
        self._assert_ok("export_impact_csv", "text/csv")
        self._assert_ok(
            "export_impact_excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self._assert_ok("export_impact_pdf", "application/pdf")

    def test_course_exports_contain_fixture_course(self):
        response = self.client.get(reverse("reports:export_course_csv"))
        self.assertIn("Export Course", response.content.decode())
        self._assert_ok(
            "export_course_excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self._assert_ok("export_course_pdf", "application/pdf")

    def test_student_exports_contain_fixture_student(self):
        response = self.client.get(reverse("reports:export_student_csv"))
        self.assertIn("Export Student", response.content.decode())
        self._assert_ok(
            "export_student_excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self._assert_ok("export_student_pdf", "application/pdf")

    def test_employment_exports_never_include_salary_or_evidence(self):
        response = self.client.get(reverse("reports:export_employment_csv"))
        content = response.content.decode()
        self.assertIn("Employed Full-Time", content)
        self.assertNotIn("salary", content.lower())
        self._assert_ok(
            "export_employment_excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self._assert_ok("export_employment_pdf", "application/pdf")

    def test_student_cannot_access_any_export(self):
        self.client.logout()
        User.objects.create_user(
            email="exportstudentuser@example.com", password="pass1234", full_name="Student",
            role=User.Role.STUDENT, status=User.Status.ACTIVE,
        )
        self.client.login(username="exportstudentuser@example.com", password="pass1234")
        for url_name in ["export_course_csv", "export_student_excel", "export_employment_pdf"]:
            response = self.client.get(reverse(f"reports:{url_name}"))
            self.assertEqual(response.status_code, 403, url_name)


class FidImpactReportExportTests(TestCase):
    """16.P — the funder-facing Impact Evidence Report exports."""

    def setUp(self):
        from apps.certificates.models import Certificate
        from apps.cohorts.models import Cohort, CohortMembership
        from apps.enrollments.models import Enrollment

        self.admin = User.objects.create_user(
            email="fidexportadmin@example.com", password="pass1234", full_name="Fid Export Admin",
            role=User.Role.ADMIN, status=User.Status.ACTIVE, is_staff=True,
        )
        self.client.login(username="fidexportadmin@example.com", password="pass1234")

        self.student = User.objects.create_user(
            email="fidexportstudent@example.com", password="pass1234", full_name="Fid Export Student",
            role=User.Role.STUDENT, status=User.Status.ACTIVE,
        )
        from apps.courses.models import Course

        self.course = Course.objects.create(
            title="Fid Export Course", slug="fid-export-course", short_description="x", description="x",
            status=Course.Status.PUBLISHED,
        )
        Enrollment.objects.create(student=self.student, course=self.course, status=Enrollment.Status.COMPLETED)
        Certificate.objects.create(student=self.student, course=self.course, status=Certificate.Status.ISSUED)
        cohort = Cohort.objects.create(name="Fid Export Cohort")
        CohortMembership.objects.create(cohort=cohort, student=self.student)

    def test_pdf_export_returns_valid_pdf(self):
        response = self.client.get(reverse("reports:export_fid_pdf"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertGreater(len(response.content), 0)

    def test_excel_export_returns_valid_workbook_with_all_sheets(self):
        import io

        from openpyxl import load_workbook

        response = self.client.get(reverse("reports:export_fid_excel"))
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(io.BytesIO(response.content))
        self.assertEqual(
            set(wb.sheetnames),
            {"Reach", "Learning Outcomes", "Employment Outcomes", "Field Impact", "Cohort Breakdown", "Top Courses"},
        )

    def test_prepared_for_query_param_appears_in_pdf(self):
        response = self.client.get(reverse("reports:export_fid_pdf"), {"prepared_for": "Test Funder Org"})
        self.assertEqual(response.status_code, 200)

    def test_malformed_prepared_for_does_not_crash_pdf_generation(self):
        response = self.client.get(reverse("reports:export_fid_pdf"), {"prepared_for": "<b><unclosed"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")

    def test_student_cannot_access_fid_exports(self):
        self.client.logout()
        User.objects.create_user(
            email="fidexportstudentuser@example.com", password="pass1234", full_name="Student",
            role=User.Role.STUDENT, status=User.Status.ACTIVE,
        )
        self.client.login(username="fidexportstudentuser@example.com", password="pass1234")
        response = self.client.get(reverse("reports:export_fid_pdf"))
        self.assertEqual(response.status_code, 403)
        response = self.client.get(reverse("reports:export_fid_excel"))
        self.assertEqual(response.status_code, 403)
